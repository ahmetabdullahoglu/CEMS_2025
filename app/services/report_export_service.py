"""
CEMS Report Export Service - خدمة تصدير التقارير
===============================================
Phase 8.1: Export functionality (JSON, Excel, PDF)
"""

import json
from datetime import datetime
from decimal import Decimal
from typing import Dict, Any, List
from io import BytesIO
import tempfile

# Excel support
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF support
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)
from reportlab.pdfgen import canvas

# Template engine
from jinja2 import Template


class DecimalEncoder(json.JSONEncoder):
    """JSON encoder that handles Decimal types"""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


class ReportExportService:
    """خدمة تصدير التقارير بصيغ متعددة"""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    # ==================== JSON EXPORT ====================
    
    def export_to_json(
        self,
        report_data: Dict[str, Any],
        pretty: bool = True
    ) -> str:
        """
        تصدير إلى JSON
        Export report to JSON format
        
        Args:
            report_data: Dictionary containing report data
            pretty: Whether to format JSON with indentation
            
        Returns:
            JSON string
        """
        try:
            if pretty:
                return json.dumps(
                    report_data,
                    cls=DecimalEncoder,
                    indent=2,
                    ensure_ascii=False
                )
            else:
                return json.dumps(
                    report_data,
                    cls=DecimalEncoder,
                    ensure_ascii=False
                )
        except Exception as e:
            raise Exception(f"Failed to export to JSON: {str(e)}")
    
    # ==================== EXCEL EXPORT ====================
    
    def export_to_excel(
        self,
        report_data: Dict[str, Any],
        template: str = 'standard'
    ) -> BytesIO:
        """
        تصدير إلى Excel
        Export report to Excel format with styling
        
        Args:
            report_data: Dictionary containing report data
            template: Template type (standard, financial, summary)
            
        Returns:
            BytesIO object containing Excel file
        """
        try:
            wb = Workbook()
            ws = wb.active
            
            # Apply template-specific styling
            if template == 'financial':
                self._apply_financial_template(ws, report_data)
            elif template == 'summary':
                self._apply_summary_template(ws, report_data)
            else:
                self._apply_standard_template(ws, report_data)
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file
            
        except Exception as e:
            raise Exception(f"Failed to export to Excel: {str(e)}")
    
    def _apply_standard_template(self, ws, report_data: Dict[str, Any]):
        """Apply standard Excel template"""
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Title
        ws['A1'] = report_data.get('title', 'Report')
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        ws[f'A{row}'] = 'Generated At:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        row += 1
        if 'date_range' in report_data:
            ws[f'A{row}'] = 'Period:'
            date_range = report_data['date_range']
            ws[f'B{row}'] = f"{date_range.get('start', '')} to {date_range.get('end', '')}"
        
        # Main data
        row += 2
        
        # If report has summary section
        if 'summary' in report_data:
            ws[f'A{row}'] = 'Summary'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:B{row}')
            
            row += 1
            summary = report_data['summary']
            for key, value in summary.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
        
        # If report has tabular data
        if 'data' in report_data or 'transactions' in report_data or 'branches' in report_data:
            data_key = 'data' if 'data' in report_data else ('transactions' if 'transactions' in report_data else 'branches')
            data = report_data[data_key]
            
            if data and isinstance(data, list):
                # Headers
                headers = list(data[0].keys()) if data else []
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header.replace('_', ' ').title()
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Data rows
                for item in data:
                    row += 1
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col)
                        value = item.get(header, '')
                        
                        # Format numbers
                        if isinstance(value, (int, float)):
                            cell.value = value
                            cell.number_format = '#,##0.00'
                        else:
                            cell.value = str(value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_financial_template(self, ws, report_data: Dict[str, Any]):
        """Apply financial-specific Excel template"""
        # Enhanced styling for financial reports
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        # Title with company info
        ws['A1'] = 'CEMS - Currency Exchange Management System'
        ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
        ws.merge_cells('A1:F1')
        
        ws['A2'] = report_data.get('title', 'Financial Report')
        ws['A2'].font = Font(size=16, bold=True)
        ws.merge_cells('A2:F2')
        
        # Financial summary with borders
        row = 4
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if 'summary' in report_data:
            summary = report_data['summary']
            
            # Summary box
            for key, value in summary.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = thin_border
                
                ws[f'B{row}'] = value
                ws[f'B{row}'].border = thin_border
                
                # Format currency values
                if 'revenue' in key.lower() or 'commission' in key.lower() or 'amount' in key.lower():
                    ws[f'B{row}'].number_format = '$#,##0.00'
                
                row += 1
        
        # Detailed data table
        if 'data' in report_data:
            row += 2
            data = report_data['data']
            
            if data and isinstance(data, list):
                headers = list(data[0].keys())
                
                # Header row with styling
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header.replace('_', ' ').title()
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                
                # Data rows
                for item in data:
                    row += 1
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col)
                        value = item.get(header, '')
                        cell.value = value
                        cell.border = thin_border
                        
                        # Format financial columns
                        if any(word in header.lower() for word in ['amount', 'revenue', 'commission', 'balance']):
                            if isinstance(value, (int, float)):
                                cell.number_format = '$#,##0.00'
        
        # Auto-adjust columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _apply_summary_template(self, ws, report_data: Dict[str, Any]):
        """Apply summary-focused Excel template"""
        # Simplified template for executive summaries
        title_font = Font(size=18, bold=True, color="203764")
        section_font = Font(size=12, bold=True, color="366092")
        
        # Title
        ws['A1'] = report_data.get('title', 'Executive Summary')
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Key metrics in a clean format
        if 'summary' in report_data:
            ws[f'A{row}'] = 'Key Metrics'
            ws[f'A{row}'].font = section_font
            row += 1
            
            summary = report_data['summary']
            for key, value in summary.items():
                ws[f'A{row}'] = f"• {key.replace('_', ' ').title()}"
                ws[f'B{row}'] = value
                
                if isinstance(value, (int, float)):
                    ws[f'B{row}'].number_format = '#,##0.00'
                
                row += 1
        
        # Charts data (if available)
        if 'charts' in report_data:
            row += 2
            ws[f'A{row}'] = 'Trend Data'
            ws[f'A{row}'].font = section_font
            row += 1
            
            # Simple table for chart data
            # This would be enhanced with actual chart objects in production
    
    # ==================== PDF EXPORT ====================
    
    def export_to_pdf(
        self,
        report_data: Dict[str, Any],
        template: str = 'standard'
    ) -> BytesIO:
        """
        تصدير إلى PDF
        Export report to PDF format
        
        Args:
            report_data: Dictionary containing report data
            template: Template type (standard, financial, summary)
            
        Returns:
            BytesIO object containing PDF file
        """
        try:
            pdf_file = BytesIO()
            
            # Create PDF document
            doc = SimpleDocTemplate(
                pdf_file,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            # Container for PDF elements
            story = []
            
            # Get styles
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12
            )
            
            # Add title
            title = report_data.get('title', 'Report')
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 12))
            
            # Add metadata
            generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            story.append(Paragraph(f"<b>Generated:</b> {generated_at}", styles['Normal']))
            
            if 'date_range' in report_data:
                date_range = report_data['date_range']
                period = f"{date_range.get('start', '')} to {date_range.get('end', '')}"
                story.append(Paragraph(f"<b>Period:</b> {period}", styles['Normal']))
            
            story.append(Spacer(1, 20))
            
            # Add summary section
            if 'summary' in report_data:
                story.append(Paragraph("Summary", heading_style))
                
                summary = report_data['summary']
                summary_data = [[key.replace('_', ' ').title(), str(value)] 
                               for key, value in summary.items()]
                
                summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.lightgrey, colors.white])
                ]))
                
                story.append(summary_table)
                story.append(Spacer(1, 20))
            
            # Add detailed data
            if 'data' in report_data or 'transactions' in report_data or 'branches' in report_data:
                data_key = 'data' if 'data' in report_data else ('transactions' if 'transactions' in report_data else 'branches')
                data = report_data[data_key]
                
                if data and isinstance(data, list):
                    story.append(Paragraph("Detailed Data", heading_style))
                    
                    # Prepare table data
                    headers = list(data[0].keys()) if data else []
                    table_data = [[h.replace('_', ' ').title() for h in headers]]
                    
                    # Limit rows for PDF (prevent huge files)
                    max_rows = 50
                    for item in data[:max_rows]:
                        row = []
                        for header in headers:
                            value = item.get(header, '')
                            if isinstance(value, (int, float)):
                                row.append(f"{value:,.2f}")
                            else:
                                row.append(str(value)[:30])  # Truncate long text
                        table_data.append(row)
                    
                    # Create table
                    col_widths = [1.5*inch] * len(headers)
                    detail_table = Table(table_data, colWidths=col_widths)
                    detail_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                    ]))
                    
                    story.append(detail_table)
                    
                    if len(data) > max_rows:
                        story.append(Spacer(1, 12))
                        story.append(Paragraph(
                            f"<i>Showing {max_rows} of {len(data)} records</i>",
                            styles['Normal']
                        ))
            
            # Build PDF
            doc.build(story)
            pdf_file.seek(0)
            
            return pdf_file
            
        except Exception as e:
            raise Exception(f"Failed to export to PDF: {str(e)}")
    
    # ==================== TEMPLATE-BASED EXPORT ====================
    
    def export_with_html_template(
        self,
        report_data: Dict[str, Any],
        html_template: str
    ) -> str:
        """
        تصدير باستخدام قالب HTML
        Export using Jinja2 HTML template
        
        Args:
            report_data: Report data
            html_template: HTML template string with Jinja2 syntax
            
        Returns:
            Rendered HTML string
        """
        try:
            template = Template(html_template)
            
            # Add helper functions to template context
            context = {
                **report_data,
                'now': datetime.now(),
                'format_currency': lambda x: f"${x:,.2f}" if isinstance(x, (int, float)) else x,
                'format_percent': lambda x: f"{x:.2f}%" if isinstance(x, (int, float)) else x
            }
            
            return template.render(**context)
            
        except Exception as e:
            raise Exception(f"Failed to render HTML template: {str(e)}")


# ==================== EXAMPLE TEMPLATES ====================

STANDARD_HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>{{ title }}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1 { color: #1F4E78; }
        h2 { color: #366092; margin-top: 30px; }
        table { border-collapse: collapse; width: 100%; margin: 20px 0; }
        th { background-color: #366092; color: white; padding: 10px; text-align: left; }
        td { padding: 8px; border: 1px solid #ddd; }
        tr:nth-child(even) { background-color: #f2f2f2; }
        .summary { background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0; }
        .footer { margin-top: 40px; font-size: 12px; color: #666; }
    </style>
</head>
<body>
    <h1>{{ title }}</h1>
    <p><strong>Generated:</strong> {{ now.strftime('%Y-%m-%d %H:%M:%S') }}</p>
    
    {% if date_range %}
    <p><strong>Period:</strong> {{ date_range.start }} to {{ date_range.end }}</p>
    {% endif %}
    
    {% if summary %}
    <div class="summary">
        <h2>Summary</h2>
        <table>
            {% for key, value in summary.items() %}
            <tr>
                <td><strong>{{ key.replace('_', ' ').title() }}</strong></td>
                <td>{{ format_currency(value) if 'revenue' in key or 'amount' in key else value }}</td>
            </tr>
            {% endfor %}
        </table>
    </div>
    {% endif %}
    
    {% if data %}
    <h2>Detailed Data</h2>
    <table>
        <thead>
            <tr>
                {% for header in data[0].keys() %}
                <th>{{ header.replace('_', ' ').title() }}</th>
                {% endfor %}
            </tr>
        </thead>
        <tbody>
            {% for item in data %}
            <tr>
                {% for value in item.values() %}
                <td>{{ value }}</td>
                {% endfor %}
            </tr>
            {% endfor %}
        </tbody>
    </table>
    {% endif %}
    
    <div class="footer">
        <p>CEMS - Currency Exchange Management System © {{ now.year }}</p>
    </div>
</body>
</html>
"""


# ==================== EXPORT HELPERS ====================

def save_export(content: BytesIO, filename: str, output_dir: str = '/tmp') -> str:
    """
    Save exported file to disk
    
    Args:
        content: BytesIO content
        filename: Target filename
        output_dir: Output directory
        
    Returns:
        Full path to saved file
    """
    import os
    
    filepath = os.path.join(output_dir, filename)
    
    with open(filepath, 'wb') as f:
        f.write(content.getvalue())
    
    return filepath


def get_export_filename(report_type: str, format: str, timestamp: bool = True) -> str:
    """
    Generate standardized export filename
    
    Args:
        report_type: Type of report (e.g., 'daily_summary')
        format: Export format (json, xlsx, pdf)
        timestamp: Whether to include timestamp
        
    Returns:
        Formatted filename
    """
    filename = f"CEMS_{report_type}"
    
    if timestamp:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename += f"_{ts}"
    
    filename += f".{format}"
    
    return filename
